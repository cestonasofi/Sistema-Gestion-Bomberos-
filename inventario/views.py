from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count
from django.core.exceptions import PermissionDenied
from django.contrib import messages
from django.core.paginator import Paginator
from django.utils import timezone

from .forms import HerramientaForm, NotaHerramientaForm, EventoCalendarioForm, ParteAsignarForm, UnidadMovilForm
from .models import Herramienta, UsuarioBombero, NotaHerramienta, RegistroActividad, EventoCalendario, ParteGuardia, ParteDetalle, Sitio, UnidadMovil

def solo_jefes(view_func):
    def wrap(request, *args, **kwargs):
        if request.user.is_authenticated and (request.user.rol == 'jefe' or request.user.is_superuser):
            return view_func(request, *args, **kwargs)
        else:
            raise PermissionDenied
    return wrap

def requiere_aprobacion(view_func):
    def wrap(request, *args, **kwargs):
        if request.user.is_authenticated and not request.user.is_approved:
            return render(request, 'inventario/esperando_aprobacion.html')
        return view_func(request, *args, **kwargs)
    return wrap


@login_required
@requiere_aprobacion
def home_inventario(request):
    query = request.GET.get('buscar', '').strip()
    
    if query:
        herramientas = Herramienta.objects.filter(
            Q(nombre__icontains=query) |
            Q(codigo__icontains=query) |
            Q(estado__icontains=query) |
            Q(disponibilidad__icontains=query) |
            Q(ubicacion__icontains=query)
        )
        total_encontrados = herramientas.count()
        
        # Mapeo de nombres legibles para unidades
        unidades_map = {
            'estructural': 'Unidad Estructural',
            'rescate': 'Unidad de Rescate',
            'forestal': 'Unidad Forestal',
            'cisterna': 'Unidad Cisterna / Hidrante',
            'deposito': 'Depósito Central'
        }
        
        unidades_set = set()
        for item in herramientas:
            unidades_set.add(unidades_map.get(item.ubicacion, item.ubicacion))
        
        unidades_encontradas = list(unidades_set)
    else:
        herramientas = Herramienta.objects.all()
        total_encontrados = herramientas.count()
        unidades_encontradas = []

    # Prefetch de notas para mostrar en cada herramienta
    herramientas = herramientas.prefetch_related('notas__usuario').order_by('codigo')

    paginator = Paginator(herramientas, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'herramientas': page_obj,
        'page_obj': page_obj,
        'query': query,
        'total_encontrados': total_encontrados,
        'unidades_encontradas': unidades_encontradas,
        'nota_form': NotaHerramientaForm(),
    }
    return render(request, 'inventario/index.html', context)


@login_required
@requiere_aprobacion
@solo_jefes
def agregar_herramienta(request):
    if request.method == 'POST':
        form = HerramientaForm(request.POST)
        if form.is_valid():
            h = form.save()
            RegistroActividad.objects.create(usuario=request.user, accion='crear', herramienta_codigo=h.codigo, herramienta_nombre=h.nombre, detalle=f'Creada por {request.user.username}')
            messages.success(request, 'Herramienta agregada con éxito al inventario.')
            return redirect('home')
    else:
        ultimo = Herramienta.objects.order_by('-codigo').values_list('codigo', flat=True).first()
        if ultimo and ultimo.isdigit():
            siguiente = str(int(ultimo) + 1).zfill(3)
        else:
            siguiente = '001'
        form = HerramientaForm(initial={'codigo': siguiente})
    return render(request, 'inventario/form_herramienta.html', {'form': form})


@login_required
@requiere_aprobacion
@solo_jefes
def editar_herramienta(request, id):
    herramienta = get_object_or_404(Herramienta, id=id)
    if request.method == 'POST':
        form = HerramientaForm(request.POST, instance=herramienta)
        if form.is_valid():
            h = form.save()
            RegistroActividad.objects.create(usuario=request.user, accion='editar', herramienta_codigo=h.codigo, herramienta_nombre=h.nombre, detalle=f'Editada por {request.user.username}')
            messages.success(request, f'Herramienta {herramienta.codigo} actualizada.')
            return redirect('home')
    else:
        form = HerramientaForm(instance=herramienta)
    return render(request, 'inventario/form_herramienta.html', {'form': form, 'herramienta': herramienta})


@login_required
@requiere_aprobacion
@solo_jefes
def eliminar_herramienta(request, id):
    herramienta = get_object_or_404(Herramienta, id=id)
    if request.method == 'POST':
        RegistroActividad.objects.create(usuario=request.user, accion='eliminar', herramienta_codigo=herramienta.codigo, herramienta_nombre=herramienta.nombre, detalle=f'Eliminada por {request.user.username}')
        herramienta.delete()
        messages.success(request, 'Herramienta eliminada del inventario.')
        return redirect('home')
    return render(request, 'inventario/confirmar_eliminar.html', {'herramienta': herramienta})


@login_required
@requiere_aprobacion
def agregar_nota(request, id):
    herramienta = get_object_or_404(Herramienta, id=id)
    if request.method == 'POST':
        form = NotaHerramientaForm(request.POST)
        if form.is_valid():
            nota = form.save(commit=False)
            nota.herramienta = herramienta
            nota.usuario = request.user
            nota.save()
            messages.success(request, f'Novedad registrada para {herramienta.nombre}.')
    return redirect('home')


@login_required
@requiere_aprobacion
def eliminar_nota(request, id):
    nota = get_object_or_404(NotaHerramienta, id=id)
    if request.method == 'POST':
        herramienta_id = nota.herramienta.id
        nota.delete()
        messages.success(request, 'Novedad eliminada correctamente.')
        return redirect('home')
    return redirect('home')


@login_required
@requiere_aprobacion
@solo_jefes
def aprobar_bombero(request, id):
    usuario = get_object_or_404(UsuarioBombero, id=id)
    usuario.is_approved = True
    usuario.save()
    messages.success(request, f'El bombero {usuario.username} ha sido aprobado correctamente.')
    return redirect('configuracion')


@login_required
@requiere_aprobacion
def configuracion(request):
    total = Herramienta.objects.count()
    disponibles = Herramienta.objects.filter(disponibilidad='Disponible').count()
    en_stock = Herramienta.objects.filter(disponibilidad='En Stock').count()
    no_disponibles = Herramienta.objects.filter(disponibilidad='No Disponible').count()
    en_uso = Herramienta.objects.filter(disponibilidad='En Uso').count()
    
    por_ubicacion = {
        'estructural': Herramienta.objects.filter(ubicacion='estructural').count(),
        'rescate': Herramienta.objects.filter(ubicacion='rescate').count(),
        'forestal': Herramienta.objects.filter(ubicacion='forestal').count(),
        'cisterna': Herramienta.objects.filter(ubicacion='cisterna').count(),
        'deposito': Herramienta.objects.filter(ubicacion='deposito').count(),
    }

    bomberos_pendientes = UsuarioBombero.objects.filter(is_approved=False)
    todos_bomberos = UsuarioBombero.objects.all().order_by('-date_joined')
    actividad_reciente = RegistroActividad.objects.all()[:10]
    eventos = EventoCalendario.objects.all()
    
    context = {
        'total': total,
        'disponibles': disponibles,
        'en_stock': en_stock,
        'no_disponibles': no_disponibles,
        'en_uso': en_uso,
        'por_ubicacion': por_ubicacion,
        'bomberos_pendientes': bomberos_pendientes,
        'todos_bomberos': todos_bomberos,
        'actividad_reciente': actividad_reciente,
        'eventos': eventos,
    }
    return render(request, 'inventario/configuracion.html', context)


@login_required
@requiere_aprobacion
@solo_jefes
def agregar_evento(request):
    if request.method == 'POST':
        form = EventoCalendarioForm(request.POST)
        if form.is_valid():
            evento = form.save(commit=False)
            evento.usuario = request.user
            evento.save()
            messages.success(request, 'Evento agregado al calendario.')
            return redirect('configuracion')
    else:
        form = EventoCalendarioForm()
    return render(request, 'inventario/form_evento.html', {'form': form})


@login_required
@requiere_aprobacion
@solo_jefes
def eliminar_evento(request, id):
    evento = get_object_or_404(EventoCalendario, id=id)
    if request.method == 'POST':
        evento.delete()
        messages.success(request, 'Evento eliminado del calendario.')
        return redirect('configuracion')
    return redirect('configuracion')


@login_required
@requiere_aprobacion
def partes_guardia(request):
    es_jefe = request.user.rol == 'jefe' or request.user.is_superuser

    if es_jefe:
        partes = ParteGuardia.objects.select_related('bombero').all()
    else:
        partes = ParteGuardia.objects.select_related('bombero').filter(bombero=request.user)

    # Resumen por unidad para el jefe
    unidades_map = dict(Herramienta.UBICACION_CHOICES)
    resumen_unidades = []
    if es_jefe:
        for clave, nombre in unidades_map.items():
            resumen_unidades.append({
                'clave': clave,
                'nombre': nombre,
                'pendientes': ParteGuardia.objects.filter(unidad=clave, estado='pendiente').count(),
                'completados': ParteGuardia.objects.filter(unidad=clave, estado='completado').count(),
            })

    context = {
        'partes': partes,
        'resumen_unidades': resumen_unidades,
        'es_jefe': es_jefe,
    }
    return render(request, 'inventario/partes.html', context)


@login_required
@requiere_aprobacion
@solo_jefes
def asignar_parte(request):
    if request.method == 'POST':
        form = ParteAsignarForm(request.POST)
        if form.is_valid():
            parte = form.save()
            messages.success(
                request,
                f'Parte asignado a {parte.bombero.username} para el {parte.fecha} ({parte.get_turno_display()}) - {parte.get_unidad_display()}.'
            )
            return redirect('partes_guardia')
    else:
        form = ParteAsignarForm(initial={'fecha': timezone.now().date()})
    return render(request, 'inventario/form_asignar_parte.html', {'form': form})


def puede_ver_parte(request, parte):
    return request.user.rol == 'jefe' or request.user.is_superuser or parte.bombero == request.user


@login_required
@requiere_aprobacion
def completar_parte(request, id):
    parte = get_object_or_404(ParteGuardia, id=id)
    if not puede_ver_parte(request, parte):
        raise PermissionDenied

    herramientas = Herramienta.objects.filter(ubicacion=parte.unidad).order_by('codigo')

    if request.method == 'POST':
        # Evitar re-enviar un parte ya completado
        if parte.estado == 'completado':
            messages.warning(request, 'Este parte ya fue completado.')
            return redirect('partes_guardia')

        for h in herramientas:
            estado = request.POST.get(f'estado_{h.id}')
            if estado not in ('Disponible', 'En Uso', 'No Disponible'):
                continue
            observacion = request.POST.get(f'obs_{h.id}', '').strip()
            ParteDetalle.objects.create(
                parte=parte,
                herramienta=h,
                estado_reporte=estado,
                observacion=observacion,
            )
            # Actualizar la herramienta en el inventario
            h.disponibilidad = estado
            if estado == 'Disponible':
                h.estado = 'Funcional'
            elif estado == 'En Uso':
                h.estado = 'En uso'
            else:
                h.estado = 'No funcional'
            h.save()
            # Registrar novedad automática si la herramienta quedó no disponible
            if estado == 'No Disponible' and observacion:
                NotaHerramienta.objects.create(
                    herramienta=h,
                    usuario=request.user,
                    detalle=f'[Parte {parte.fecha} {parte.get_turno_display()}] {observacion}',
                )

        parte.estado = 'completado'
        parte.completado = timezone.now()
        parte.save()
        RegistroActividad.objects.create(
            usuario=request.user,
            accion='editar',
            herramienta_codigo='PARTE',
            herramienta_nombre=f'Parte {parte.get_unidad_display()} {parte.fecha}',
            detalle=f'Completado por {request.user.username}',
        )
        messages.success(request, f'Parte de {parte.get_unidad_display()} ({parte.fecha}) completado. El inventario fue actualizado.')
        return redirect('partes_guardia')

    return render(request, 'inventario/completar_parte.html', {'parte': parte, 'herramientas': herramientas})


@login_required
@requiere_aprobacion
def ver_parte(request, id):
    parte = get_object_or_404(ParteGuardia, id=id)
    if not puede_ver_parte(request, parte):
        raise PermissionDenied
    return render(request, 'inventario/ver_parte.html', {'parte': parte})


@login_required
@requiere_aprobacion
@solo_jefes
def eliminar_parte(request, id):
    parte = get_object_or_404(ParteGuardia, id=id)
    if request.method == 'POST':
        parte.delete()
        messages.success(request, 'Parte eliminado.')
        return redirect('partes_guardia')
    return redirect('partes_guardia')


@login_required
@requiere_aprobacion
@solo_jefes
def vaciar_actividad(request):
    if request.method == 'POST':
        RegistroActividad.objects.all().delete()
        messages.success(request, 'Registro de actividad eliminado por completo.')
    return redirect('configuracion')


# ============ ACCESO POR ROL ============

@login_required
def panel_principal(request):
    # Al ingresar, todos los roles van directo al inventario
    if not request.user.is_approved:
        return render(request, 'inventario/esperando_aprobacion.html')
    return redirect('home')


@login_required
@requiere_aprobacion
def panel_bombero(request):
    # Acceso exclusivo para bomberos rescatistas
    if request.user.rol != 'rescatista':
        raise PermissionDenied
    partes_pendientes = ParteGuardia.objects.filter(
        bombero=request.user, estado='pendiente'
    ).select_related().order_by('fecha', 'turno')

    completados = ParteGuardia.objects.filter(
        bombero=request.user, estado='completado'
    ).count()

    total_herramientas = Herramienta.objects.count()
    disponibles = Herramienta.objects.filter(disponibilidad='Disponible').count()
    no_disponibles = Herramienta.objects.filter(disponibilidad='No Disponible').count()

    unidades = UnidadMovil.objects.select_related('sitio').order_by('sitio__nombre', 'nombre')
    unidades_servicio = UnidadMovil.objects.filter(estado_operativo='servicio').count()

    context = {
        'partes_pendientes': partes_pendientes,
        'completados': completados,
        'total_herramientas': total_herramientas,
        'disponibles': disponibles,
        'no_disponibles': no_disponibles,
        'unidades': unidades,
        'unidades_servicio': unidades_servicio,
    }
    return render(request, 'inventario/panel_bombero.html', context)


# ============ CATÁLOGO DE UNIDADES MÓVILES ============

@login_required
@requiere_aprobacion
def unidades_lista(request):
    sitio_filter = request.GET.get('sitio', '')
    sitios = Sitio.objects.prefetch_related('unidades').all()

    if sitio_filter:
        sitios = sitios.filter(id=sitio_filter)

    context = {
        'sitios': sitios,
        'sitio_filter': sitio_filter,
        'es_jefe': request.user.rol == 'jefe' or request.user.is_superuser,
    }
    return render(request, 'inventario/unidades.html', context)


@login_required
@requiere_aprobacion
@solo_jefes
def agregar_unidad(request):
    if request.method == 'POST':
        form = UnidadMovilForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Unidad agregada correctamente.')
            return redirect('unidades_lista')
    else:
        form = UnidadMovilForm()
    return render(request, 'inventario/form_unidad.html', {'form': form, 'titulo': 'Agregar Unidad'})


@login_required
@requiere_aprobacion
@solo_jefes
def editar_unidad(request, id):
    unidad = get_object_or_404(UnidadMovil, id=id)
    if request.method == 'POST':
        form = UnidadMovilForm(request.POST, instance=unidad)
        if form.is_valid():
            form.save()
            messages.success(request, f'Unidad {unidad.nombre} actualizada.')
            return redirect('unidades_lista')
    else:
        form = UnidadMovilForm(instance=unidad)
    return render(request, 'inventario/form_unidad.html', {'form': form, 'unidad': unidad, 'titulo': f'Editar {unidad.nombre}'})


@login_required
@requiere_aprobacion
@solo_jefes
def eliminar_unidad(request, id):
    unidad = get_object_or_404(UnidadMovil, id=id)
    if request.method == 'POST':
        nombre = unidad.nombre
        unidad.delete()
        messages.success(request, f'Unidad {nombre} eliminada.')
    return redirect('unidades_lista')


# ---------------------------------------------------------------------------
# Exportación y reportes
# ---------------------------------------------------------------------------

def _filas_inventario():
    """Devuelve las herramientas ordenadas con columnas legibles para exportar."""
    filas = []
    for h in Herramienta.objects.select_related('unidad').order_by('codigo'):
        filas.append([
            h.codigo,
            h.nombre,
            h.get_ubicacion_display(),
            h.unidad.nombre if h.unidad else '',
            h.estado,
            h.disponibilidad,
        ])
    return filas


@login_required
@requiere_aprobacion
def exportar_inventario_csv(request):
    import csv

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="inventario.csv"'
    response.write('\ufeff')  # BOM para que Excel muestre bien los acentos
    writer = csv.writer(response)
    writer.writerow(['Código', 'Nombre', 'Ubicación', 'Unidad', 'Estado', 'Disponibilidad'])
    for fila in _filas_inventario():
        writer.writerow(fila)
    return response


@login_required
@requiere_aprobacion
def exportar_inventario_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    headers = ['Código', 'Nombre', 'Ubicación', 'Unidad', 'Estado', 'Disponibilidad']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'
    ws.append(headers)

    for celda in ws[1]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill(start_color='C1121F', end_color='C1121F', fill_type='solid')
        celda.alignment = Alignment(horizontal='center')

    for fila in _filas_inventario():
        ws.append(fila)

    for col in ws.columns:
        ancho = max(len(str(celda.value)) for celda in col) + 4
        ws.column_dimensions[col[0].column_letter].width = min(ancho, 60)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="inventario.xlsx"'
    wb.save(response)
    return response


@login_required
@requiere_aprobacion
def reporte_inventario(request):
    from collections import OrderedDict

    herramientas = Herramienta.objects.select_related('unidad').order_by('ubicacion', 'codigo')
    orden_ubicaciones = ['estructural', 'rescate', 'forestal', 'cisterna', 'deposito']
    grupos = OrderedDict()
    for ubicacion in orden_ubicaciones:
        grupos[ubicacion] = [h for h in herramientas if h.ubicacion == ubicacion]

    context = {
        'grupos': grupos,
        'total': herramientas.count(),
        'fecha': timezone.now(),
        'usuario': request.user,
    }
    return render(request, 'inventario/reporte_inventario.html', context)


@login_required
@requiere_aprobacion
def reporte_inventario_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    estilos = getSampleStyleSheet()
    titulo_estilo = ParagraphStyle(
        'Titulo', parent=estilos['Title'], textColor=colors.HexColor('#C1121F'), fontSize=20
    )
    subtitulo_estilo = ParagraphStyle(
        'Subtitulo', parent=estilos['Normal'], alignment=TA_CENTER, textColor=colors.HexColor('#555555')
    )
    celda_estilo = ParagraphStyle(
        'Celda', parent=estilos['Normal'], fontSize=8, leading=10
    )

    elementos = []
    elementos.append(Paragraph('Reporte de Inventario', titulo_estilo))
    elementos.append(Paragraph(f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")} - Operador: {request.user.username}', subtitulo_estilo))
    elementos.append(Spacer(1, 8 * mm))

    encabezado = [Paragraph(f'<b>{c}</b>', celda_estilo) for c in
                  ['Código', 'Material / Herramienta', 'Unidad', 'Estado', 'Disponibilidad']]

    ubicaciones = ['estructural', 'rescate', 'forestal', 'cisterna', 'deposito']
    nombres_ubicacion = dict(Herramienta.UBICACION_CHOICES)

    for ubicacion in ubicaciones:
        filas = _filas_inventario_por_ubicacion(ubicacion)
        if not filas:
            continue
        elementos.append(Paragraph(f'<b>{nombres_ubicacion.get(ubicacion, ubicacion)}</b>', estilos['Heading3']))
        elementos.append(Spacer(1, 3 * mm))
        tabla_datos = [[Paragraph(str(valor), celda_estilo) for valor in fila] for fila in filas]
        tabla = Table([encabezado] + tabla_datos, repeatRows=1)
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C1121F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F7F7')]),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        tabla._argW = [25 * mm, 90 * mm, 55 * mm, 35 * mm, 40 * mm]
        elementos.append(tabla)
        elementos.append(Spacer(1, 8 * mm))

    doc.build(elementos)
    return response


def _filas_inventario_por_ubicacion(ubicacion):
    filas = []
    for h in Herramienta.objects.filter(ubicacion=ubicacion).select_related('unidad').order_by('codigo'):
        filas.append([
            h.codigo,
            h.nombre,
            h.unidad.nombre if h.unidad else '',
            h.estado,
            h.disponibilidad,
        ])
    return filas
